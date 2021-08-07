import json
import requests
import pytest
import openpyxl


@pytest.mark.usefixtures("api_url", "api_key")
class TestDeleteMail:

#Open file with email info
    def test_delete_mail_by_id(self, api_url, api_key):
        path = "Email_info.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet_obj = workbook.active

#Get original id from each email
        x=2
        cell_obj = sheet_obj.cell(row=x, column=5)

        while cell_obj.value is not None:
            cell_obj = sheet_obj.cell(row=x, column=5)
            if cell_obj.value == None:
                break
            else:
#Delete email based on original id
                response = requests.delete(api_url + "/api/addresses/mailsacfortesting@mailsac.com/messages/" + cell_obj.value, headers = api_key)
                response.status_code == 200
                json_response = json.loads(response.text)
                print(json_response)
            x+=1





